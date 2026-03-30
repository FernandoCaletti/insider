"""Holdings endpoints."""

import csv
import io
from datetime import date
from typing import Any

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from api.app.database import get_cursor

_EXPORT_COLUMNS = [
    ("empresa", "Empresa"),
    ("ticker", "Ticker"),
    ("data", "Data"),
    ("tipo_ativo", "Tipo de Ativo"),
    ("descricao", "Descrição"),
    ("operacao", "Operação"),
    ("quantidade", "Quantidade"),
    ("preco_unitario", "Preço Unitário"),
    ("valor_total", "Valor Total"),
    ("corretora", "Corretora"),
    ("nome_insider", "Nome do Insider"),
    ("secao", "Seção"),
    ("confianca", "Confiança"),
    ("data_referencia", "Data de Referência"),
]

router = APIRouter(prefix="/holdings", tags=["holdings"])

_VALID_INSIDER_GROUPS = [
    "Controlador",
    "Conselho de Administracao",
    "Diretoria",
    "Conselho Fiscal",
    "Orgaos Tecnicos",
    "Pessoas Ligadas",
]

_INSIDER_GROUP_LOWER = {g.lower(): g for g in _VALID_INSIDER_GROUPS}


def _validate_insider_group(value: str | None) -> str | None:
    """Validate and normalize insider_group (case-insensitive). Returns canonical value or raises 400."""
    if value is None:
        return None
    canonical = _INSIDER_GROUP_LOWER.get(value.lower())
    if canonical is None:
        raise HTTPException(
            status_code=400,
            detail=f"insider_group invalido. Valores aceitos: {', '.join(_VALID_INSIDER_GROUPS)}",
        )
    return canonical

# Allowed sort columns to prevent SQL injection
_SORT_COLUMNS = {
    "operation_date": "h.operation_date",
    "total_value": "h.total_value",
    "quantity": "h.quantity",
    "asset_type": "h.asset_type",
    "company_name": "c.name",
    "operation_type": "h.operation_type",
    "unit_price": "h.unit_price",
    "broker": "h.broker",
    "company_ticker": "c.ticker",
}


def _build_holdings_query(
    company_id: int | None,
    asset_type: str | None,
    operation_type: str | None,
    date_from: date | None,
    date_to: date | None,
    value_min: float | None,
    value_max: float | None,
    section: str | None = None,
    insider_group: str | None = None,
) -> tuple[str, list[Any], dict[str, Any]]:
    """Build WHERE clause, params, and filters_applied dict for holdings queries."""
    conditions = ["h.confidence != 'baixa'"]
    params: list[Any] = []
    filters_applied: dict[str, Any] = {}

    if company_id is not None:
        conditions.append("d.company_id = %s")
        params.append(company_id)
        filters_applied["company_id"] = company_id
    if section:
        conditions.append("h.section = %s")
        params.append(section)
        filters_applied["section"] = section
    if asset_type:
        types = [t.strip() for t in asset_type.split(",") if t.strip()]
        if len(types) == 1:
            conditions.append("h.asset_type = %s")
            params.append(types[0])
        else:
            placeholders = ", ".join(["%s"] * len(types))
            conditions.append(f"h.asset_type IN ({placeholders})")
            params.extend(types)
        filters_applied["asset_type"] = asset_type
    if operation_type:
        if operation_type == "mercado":
            conditions.append("(h.operation_type ILIKE 'Compra%%' OR h.operation_type ILIKE 'Venda%%')")
        elif operation_type == "corporativa":
            conditions.append("NOT (h.operation_type ILIKE 'Compra%%' OR h.operation_type ILIKE 'Venda%%')")
        else:
            conditions.append("h.operation_type ILIKE %s")
            params.append(f"{operation_type}%")
        filters_applied["operation_type"] = operation_type
    if date_from:
        conditions.append("h.operation_date >= %s")
        params.append(date_from)
        filters_applied["date_from"] = date_from.isoformat()
    if date_to:
        conditions.append("h.operation_date <= %s")
        params.append(date_to)
        filters_applied["date_to"] = date_to.isoformat()
    if value_min is not None:
        conditions.append("h.total_value >= %s")
        params.append(value_min)
        filters_applied["value_min"] = value_min
    if value_max is not None:
        conditions.append("h.total_value <= %s")
        params.append(value_max)
        filters_applied["value_max"] = value_max
    if insider_group is not None:
        conditions.append("LOWER(h.insider_group) = LOWER(%s)")
        params.append(insider_group)
        filters_applied["insider_group"] = insider_group

    where = "WHERE " + " AND ".join(conditions)
    return where, params, filters_applied


@router.get("")
async def list_holdings(
    company_id: int | None = None,
    asset_type: str | None = None,
    operation_type: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    value_min: float | None = None,
    value_max: float | None = None,
    section: str | None = None,
    insider_group: str | None = None,
    sort_by: str = Query("operation_date", pattern="^(operation_date|total_value|quantity|asset_type|company_name|operation_type|unit_price|broker|company_ticker)$"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
) -> dict[str, Any]:
    """List holdings with filters."""
    validated_group = _validate_insider_group(insider_group)
    where, params, filters_applied = _build_holdings_query(
        company_id, asset_type, operation_type, date_from, date_to, value_min, value_max, section,
        insider_group=validated_group,
    )
    offset = (page - 1) * per_page
    sort_col = _SORT_COLUMNS.get(sort_by, "h.operation_date")
    order = "ASC" if sort_order == "asc" else "DESC"

    with get_cursor() as cur:
        # Count total
        cur.execute(
            f"""
            SELECT COUNT(*) AS cnt
            FROM holdings h
            JOIN documents d ON d.id = h.document_id
            JOIN companies c ON c.id = d.company_id
            {where}
            """,
            params,
        )
        total = cur.fetchone()["cnt"]  # type: ignore[index]

        # Paginated results
        query_params = [*params, per_page, offset]
        cur.execute(
            f"""
            SELECT
                h.*,
                d.reference_date,
                c.id AS company_id,
                c.name AS company_name,
                c.ticker AS company_ticker
            FROM holdings h
            JOIN documents d ON d.id = h.document_id
            JOIN companies c ON c.id = d.company_id
            {where}
            ORDER BY {sort_col} {order}, h.id
            LIMIT %s OFFSET %s
            """,
            query_params,
        )
        rows = cur.fetchall()

    return {
        "data": [dict(r) for r in rows],  # type: ignore[arg-type]
        "total": total,
        "page": page,
        "per_page": per_page,
        "filters_applied": filters_applied,
    }


@router.get("/export")
async def export_holdings(
    company_id: int | None = None,
    asset_type: str | None = None,
    operation_type: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    value_min: float | None = None,
    value_max: float | None = None,
    section: str | None = None,
    insider_group: str | None = None,
    sort_by: str = Query("operation_date", pattern="^(operation_date|total_value|quantity|asset_type|company_name|operation_type|unit_price|broker|company_ticker)$"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
) -> StreamingResponse:
    """Export holdings as CSV (max 10000 records)."""
    validated_group = _validate_insider_group(insider_group)
    where, params, _ = _build_holdings_query(
        company_id, asset_type, operation_type, date_from, date_to, value_min, value_max, section,
        insider_group=validated_group,
    )
    sort_col = _SORT_COLUMNS.get(sort_by, "h.operation_date")
    order = "ASC" if sort_order == "asc" else "DESC"
    max_records = 10000

    with get_cursor() as cur:
        cur.execute(
            f"""
            SELECT
                c.name AS empresa,
                c.ticker,
                h.operation_date AS data,
                h.asset_type AS tipo_ativo,
                h.asset_description AS descricao,
                h.operation_type AS operacao,
                h.quantity AS quantidade,
                h.unit_price AS preco_unitario,
                h.total_value AS valor_total,
                h.broker AS corretora,
                h.insider_name AS nome_insider,
                h.section AS secao,
                h.confidence AS confianca,
                d.reference_date AS data_referencia
            FROM holdings h
            JOIN documents d ON d.id = h.document_id
            JOIN companies c ON c.id = d.company_id
            {where}
            ORDER BY {sort_col} {order}, h.id
            LIMIT %s
            """,
            [*params, max_records],
        )
        rows = cur.fetchall()

    # Build CSV in memory
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))  # type: ignore[arg-type]
        writer.writeheader()
        for row in rows:
            writer.writerow(dict(row))  # type: ignore[arg-type]
    else:
        output.write("empresa,ticker,data,tipo_ativo,descricao,operacao,quantidade,preco_unitario,valor_total,corretora,nome_insider,secao,confianca,data_referencia\n")

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=holdings_export.csv"},
    )


def _fetch_export_rows(
    company_id: int | None,
    asset_type: str | None,
    operation_type: str | None,
    date_from: date | None,
    date_to: date | None,
    value_min: float | None,
    value_max: float | None,
    section: str | None,
    insider_group: str | None,
    sort_by: str,
    sort_order: str,
) -> list[Any]:
    """Shared helper to fetch export rows for CSV/XLSX."""
    validated_group = _validate_insider_group(insider_group)
    where, params, _ = _build_holdings_query(
        company_id, asset_type, operation_type, date_from, date_to, value_min, value_max, section,
        insider_group=validated_group,
    )
    sort_col = _SORT_COLUMNS.get(sort_by, "h.operation_date")
    order = "ASC" if sort_order == "asc" else "DESC"
    max_records = 10000

    with get_cursor() as cur:
        cur.execute(
            f"""
            SELECT
                c.name AS empresa,
                c.ticker,
                h.operation_date AS data,
                h.asset_type AS tipo_ativo,
                h.asset_description AS descricao,
                h.operation_type AS operacao,
                h.quantity AS quantidade,
                h.unit_price AS preco_unitario,
                h.total_value AS valor_total,
                h.broker AS corretora,
                h.insider_name AS nome_insider,
                h.section AS secao,
                h.confidence AS confianca,
                d.reference_date AS data_referencia
            FROM holdings h
            JOIN documents d ON d.id = h.document_id
            JOIN companies c ON c.id = d.company_id
            {where}
            ORDER BY {sort_col} {order}, h.id
            LIMIT %s
            """,
            [*params, max_records],
        )
        return cur.fetchall()


@router.get("/export/xlsx")
async def export_holdings_xlsx(
    company_id: int | None = None,
    asset_type: str | None = None,
    operation_type: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    value_min: float | None = None,
    value_max: float | None = None,
    section: str | None = None,
    insider_group: str | None = None,
    sort_by: str = Query("operation_date", pattern="^(operation_date|total_value|quantity|asset_type|company_name|operation_type|unit_price|broker|company_ticker)$"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
) -> StreamingResponse:
    """Export holdings as formatted XLSX (max 10000 records)."""
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment, numbers  # type: ignore[import-untyped]

    rows = _fetch_export_rows(
        company_id, asset_type, operation_type, date_from, date_to,
        value_min, value_max, section, insider_group, sort_by, sort_order,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentações"  # type: ignore[union-attr]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header row
    headers = [label for _, label in _EXPORT_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    currency_cols = {8, 9}  # preco_unitario, valor_total (1-indexed)
    quantity_col = 7
    db_keys = [key for key, _ in _EXPORT_COLUMNS]

    for row_idx, row in enumerate(rows, 2):
        row_dict: dict[str, Any] = dict(row)  # type: ignore[arg-type]
        for col_idx, key in enumerate(db_keys, 1):
            val = row_dict.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)  # type: ignore[union-attr]
            if val is not None and col_idx in currency_cols:
                cell.value = float(val)  # type: ignore[assignment]
                cell.number_format = '#,##0.00'
            elif val is not None and col_idx == quantity_col:
                cell.value = float(val) if val else 0  # type: ignore[assignment]
                cell.number_format = '#,##0'
            elif hasattr(val, 'isoformat'):
                cell.value = val  # type: ignore[assignment]
                cell.number_format = numbers.FORMAT_DATE_DATETIME
            else:
                cell.value = str(val) if val is not None else ""  # type: ignore[assignment]

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))  # type: ignore[union-attr]
        for row_idx in range(2, min(len(rows) + 2, 102)):  # Sample first 100 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value  # type: ignore[union-attr]
            max_len = max(max_len, len(str(cell_val or "")))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)  # type: ignore[union-attr]

    # Freeze header row
    ws.freeze_panes = "A2"  # type: ignore[union-attr]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=holdings_export.xlsx"},
    )
